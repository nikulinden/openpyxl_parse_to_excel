import xml.etree.ElementTree as ET
import openpyxl
from openpyxl import Workbook
import os


def create_workbook(spreadsheet_name):
    """create and return an empty excel workbook
    input - the name of the workbook
    return - empty workbook """

    workbook = Workbook()
    workbook.save(filename=spreadsheet_name)
    return workbook


def get_xml_filenames(working_dir):
    """get the names of xml files
    return list of file names """

    file_names = list()
    with os.scandir(working_dir) as entries:
        for entry in entries:
            if entry.is_file and entry.name.endswith('.xml'):
                file_names.append(entry.name)

    return file_names


def parse_xml_to_dict(root):
    """parse xml weights to dictionary of dictionaries
        return: the dictionary of dictionaries created """
    feature_weights = dict()
    for weight in root.findall('weighttable'):          #traverse the xml for each weight table
        name = weight.get('name')                       #get a name of each weight table
        feature_weights[str(name)] = dict()
        counter = 0                                     #crate a dict within the main dict for each weight table
        for item in weight.findall('weightentry'):      #traverse each item of a weighttable
            count = int(item.get('count'))              #get count value
            value = item.get('value')                  #get value
            if ':' in value:
                value = value.split(':', maxsplit=1)    #format value
                value = str(value[1])
            if count == 0:                              #skip 'NONE' entries with 0 weights
                pass                                    #TODO
            elif value in feature_weights[name]:          #check if there is a duplicate value and change the name so it is unique
                value = value+str(counter)
                feature_weights[name][value] = count
                counter += 1   #increase the counter
            else:
                feature_weights[name][value] = count       #create a new entry if does not yet exist
        counter = 0
    return feature_weights


def get_xml_reels_with_weights(root):
    """get the reels from xml if the reels have weight"""
    reels = dict()
    tmp_list = list()
    tmp_list_weights = list()
    for reel in root.findall('reelstripdef'):
        reel_name = reel.get('name')
        for symbol_name in reel.findall('stop'):
            name = symbol_name.get('symbolname')
            reel_weight = symbol_name.get('weight')
            tmp_list.append(name)
            tmp_list_weights.append(int(reel_weight))
        reels[reel_name] = tmp_list.copy()
        reels[reel_name+'_weights'] = tmp_list_weights.copy()
        tmp_list.clear()
        tmp_list_weights.clear()
    return reels


def get_xml_reels(root):
    """get the reels from xml (no weights)"""
    reels = dict()
    tmp_list = list()
    for reel in root.findall('reelstripdef'):
        reel_name = reel.get('name')
        for symbol_name in reel.findall('stop'):
            name = symbol_name.get('symbolname')
            tmp_list.append(name)
        reels[reel_name] = tmp_list.copy()
        tmp_list.clear()
    return reels


def get_main_slot_combo_xml(root):
    """parse data from main_slot/freegames_slot
    input - the root of the CombinationGroups xml tag
    return value - the dictionary with the symbol-pay hash map """

    symbol_definitions = dict()
    sym_sub_dict = dict()
    for groups in root.findall('CombinationGroups'):
        for group in groups.findall('CombinationGroup'):
            for combo in group.findall('Combination'):
                name = combo.get('Id')
                award = combo.get('Award')
                sym_sub_dict[name] = int(award)
    return sym_sub_dict


def get_main_slot_symbols_xml(root):
    """parse data from main_slot/freegames_slot
    input - the root of the symboldefs xml tag
    return value - the list of dicts with the symbols hash map """

    symbol_definitions = dict()
    symbol_substitution = list()
    for symbols in root.findall('symboldefs'):
        for symbol in symbols.findall('symboldef'):
            id = int(symbol.get('id'))
            name = symbol.get('name')
            symbol_definitions[id] = name
            #print(id, symbol_definitions[id])
    return symbol_substitution


def dict_to_excel (feature_weights, spreadsheet_name, sheet_name):
    """paste the parsed xml values to excel """
    wb = openpyxl.load_workbook(spreadsheet_name)           #TODO -- need refactoring
    #check if the sheet is already created
    if sheet_name in wb.sheetnames:
        #set the var to the created sheet
        sheet = wb[sheet_name]
    else:
        #create a new sheet
        sheet = wb.create_sheet(sheet_name)
    #delete first two cols
    sheet.delete_cols(idx=1, amount=2)
    row = 1
    for key,values in feature_weights.items():
        sheet.cell(row=row, column=1, value=key)
        row += 1
        for weight in values:
            sheet.cell(row=row, column=1, value=weight)
            sheet.cell(row=row, column=2, value=values[weight])
            row += 1

    wb.save(spreadsheet_name)


def list_to_excel (reels, spreadsheet_name, sheet_name):
    """paste the parsed xml values to excel """
    wb = openpyxl.load_workbook(spreadsheet_name)
    #check if the sheet is already created
    if sheet_name in wb.sheetnames:
        #set the var to the created sheet
        sheet = wb[sheet_name]
    else:
        #create a new sheet
        sheet = wb.create_sheet(sheet_name)
    #delete first 5 col
    sheet.delete_cols(idx=1, amount=5)
    row = 1
    col = 1
    for key,values in reels.items():
        #first print the key value
        sheet.cell(row=row, column=col, value=key)
        row += 1
        for weight in values:
            #print values
            sheet.cell(row=row, column=col, value=weight)
            row += 1
        row = 1
        col += 1 #once the first dictionary is printed go to the next col
    wb.save(spreadsheet_name)


def main_dict_to_excel (feature_weights, spreadsheet_name, sheet_name):
    """paste the parsed xml values to excel """
    wb = openpyxl.load_workbook(spreadsheet_name)           #TODO -- need refactoring
    #check if the sheet is already created
    if sheet_name in wb.sheetnames:
        #set the var to the created sheet
        sheet = wb[sheet_name]
    else:
        #create a new sheet
        sheet = wb.create_sheet(sheet_name)
    #delete first two cols
    sheet.delete_cols(idx=1, amount=2)
    row = 1
    for key,values in feature_weights.items():
        sheet.cell(row=row, column=1, value=key)
        sheet.cell(row=row, column=2, value=values)
        row += 1

    wb.save(spreadsheet_name)
